import os
import json
from collections import defaultdict
from itertools import combinations
import openpyxl
from openpyxl import Workbook
from datetime import datetime

output_dir = r'C:\Users\JuanNava\PycharmProjects\PythonProject\Optus_Topology_Creator\Output\stp'


def save_analysis_to_excel(device_vlans, device_ranges, output_dir):
    wb = Workbook()
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Switch", "VLANs", "Ranges"])
    for dev in device_vlans:
        vlans = ', '.join(sorted(device_vlans[dev]))
        ranges = ', '.join(device_ranges[dev])
        ws1.append([dev, vlans, ranges])

    # Sheet 2: Exact VLANs
    ws2 = wb.create_sheet("Exact")
    ws2.append(["Group", "Switches", "VLANs", "Ranges"])
    vlan_hash = defaultdict(list)
    for dev, vlans in device_vlans.items():
        key = (frozenset(vlans), tuple(device_ranges[dev]))
        vlan_hash[key].append(dev)
    group_num = 1
    for (vlans, ranges), group in vlan_hash.items():
        if len(group) > 1:
            ws2.append([group_num, ', '.join(group), ', '.join(sorted(vlans)), ', '.join(ranges)])
            group_num += 1

    # Sheet 3: Similar VLANs
    ws3 = wb.create_sheet("Similar")
    ws3.append(["Switch 1", "Switch 2", "Similarity", "VLANs in Common"])
    checked = set()
    for dev1, dev2 in combinations(device_vlans.keys(), 2):
        set1, set2 = device_vlans[dev1], device_vlans[dev2]
        ranges1, ranges2 = device_ranges[dev1], device_ranges[dev2]
        if ranges1 and ranges2 and ranges1 == ranges2:
            continue
        if set1 and set2:
            intersection = len(set1 & set2)
            union = len(set1 | set2)
            if union > 0 and intersection / union > 0.8 and set1 != set2:
                pair = tuple(sorted([dev1, dev2]))
                if pair not in checked:
                    ws3.append([dev1, dev2, f"{intersection/union:.2f}", ', '.join(sorted(set1 & set2))])
                    checked.add(pair)

    # Sheet 4: Migration Order
    ws4 = wb.create_sheet("Migration")
    ws4.append(["Order", "Switches", "VLANs", "Ranges", "Group Size"])
    group_vlans = {}
    for dev, vlans in device_vlans.items():
        key = (frozenset(vlans), tuple(device_ranges[dev]))
        group_vlans[key] = (vlans, device_ranges[dev])
    def vlan_count(key):
        vlans, ranges = group_vlans[key]
        return len(vlans) + len(ranges)
    migration_keys = sorted(vlan_hash.keys(), key=vlan_count)
    for idx, key in enumerate(migration_keys, 1):
        group = vlan_hash[key]
        vlans, ranges = group_vlans[key]
        ws4.append([idx, ', '.join(group), ', '.join(sorted(vlans)), ', '.join(ranges), len(group)])

    # Save file
    dt_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"stp_migration_analysis_{dt_str}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"Excel file saved to {filepath}")

def get_latest_json_file(directory):
    files = [f for f in os.listdir(directory) if f.endswith('.json')]
    files.sort(reverse=True)
    return os.path.join(directory, files[0]) if files else None

def parse_vlans(mapped):
    vlans = set()
    ranges = []
    for part in mapped.split(','):
        part = part.strip()
        if '-' in part:
            ranges.append(part)
        else:
            vlans.add(part)
    return vlans, ranges

def get_device_vlans(data):
    device_vlans = {}
    device_ranges = {}
    for device_name, info in data.get('device', {}).items():
        vlans = set()
        ranges = []
        vlan_data = info.get('vlan_data')
        if vlan_data:
            vlans = set(vlan_data.keys())
        else:
            mst_instances = info.get('mst_instances')
            if mst_instances:
                for inst in mst_instances.values():
                    mapped = inst.get('vlans_mapped')
                    if mapped and mapped != "none":
                        v, r = parse_vlans(mapped)
                        vlans.update(v)
                        ranges.extend(r)
        device_vlans[device_name] = vlans
        device_ranges[device_name] = ranges
    return device_vlans, device_ranges

def print_summary(device_vlans, device_ranges):
    print("STP State Summary:")
    print(f"Total switches: {len(device_vlans)}")
    for dev in device_vlans:
        vlan_list = sorted(device_vlans[dev])
        range_list = device_ranges[dev]
        if range_list:
            print(f"{dev}: VLAN ranges {', '.join(range_list)}")
        else:
            print(f"{dev}: VLANs {', '.join(vlan_list)}")

    # Exact VLAN matches
    vlan_hash = defaultdict(list)
    for dev, vlans in device_vlans.items():
        key = (frozenset(vlans), tuple(device_ranges[dev]))
        vlan_hash[key].append(dev)
    print("\nSwitches with exactly the same VLANs:")
    for (vlans, ranges), group in vlan_hash.items():
        if len(group) > 1:
            if ranges:
                print(f"{', '.join(group)} (VLAN range: {', '.join(ranges)})")
            else:
                print(f"{', '.join(group)} (VLANs in common: {', '.join(sorted(vlans))})")

    # Similar VLANs (Jaccard > 0.8)
    print("\nSwitches with close to the same VLANs (Jaccard > 0.8):")
    checked = set()
    for dev1, dev2 in combinations(device_vlans.keys(), 2):
        set1, set2 = device_vlans[dev1], device_vlans[dev2]
        ranges1, ranges2 = device_ranges[dev1], device_ranges[dev2]
        if ranges1 and ranges2 and ranges1 == ranges2:
            continue  # Already handled in exact match
        if set1 and set2:
            intersection = len(set1 & set2)
            union = len(set1 | set2)
            if union > 0 and intersection / union > 0.8 and set1 != set2:
                pair = tuple(sorted([dev1, dev2]))
                if pair not in checked:
                    print(f"{dev1}, {dev2} (Similarity: {intersection/union:.2f}, VLANs in common: {', '.join(sorted(set1 & set2))})")
                    checked.add(pair)


def recommend_migration_order(device_vlans, device_ranges):
    vlan_hash = defaultdict(list)
    group_vlans = {}
    for dev, vlans in device_vlans.items():
        key = (frozenset(vlans), tuple(device_ranges[dev]))
        vlan_hash[key].append(dev)
        group_vlans[key] = (vlans, device_ranges[dev])
    # Sort by total VLAN count (individual + ranges), ascending
    def vlan_count(key):
        vlans, ranges = group_vlans[key]
        return len(vlans) + len(ranges)
    migration_keys = sorted(vlan_hash.keys(), key=vlan_count)
    print("Recommended VLAN Migration Order (least VLANs first):")
    for idx, key in enumerate(migration_keys, 1):
        group = vlan_hash[key]
        vlans, ranges = group_vlans[key]
        if len(group) > 1:
            if ranges:
                print(f"Group {idx}: {', '.join(group)} (migrate together) - VLAN ranges: {', '.join(ranges)}")
            else:
                print(f"Group {idx}: {', '.join(group)} (migrate together) - VLANs: {', '.join(sorted(vlans))}")
        else:
            if ranges:
                print(f"Group {idx}: {group[0]} (single switch) - VLAN ranges: {', '.join(ranges)}")
            else:
                print(f"Group {idx}: {group[0]} (single switch) - VLANs: {', '.join(sorted(vlans))}")


def create_summary():
    json_file = get_latest_json_file(output_dir)
    if not json_file:
        print("No JSON file found.")
        return
    with open(json_file, 'r') as f:
        data = json.load(f)
    device_vlans, device_ranges = get_device_vlans(data)
    print_summary(device_vlans, device_ranges)

    print("\n\n########################################################################")
    print("####################### VLAN MIGRATION ORDER #############################")
    print("########################################################################\n")
    recommend_migration_order(device_vlans, device_ranges)

    print("")
    save_analysis_to_excel(device_vlans, device_ranges,
                           r'C:\Users\JuanNava\PycharmProjects\PythonProject\Optus_Topology_Creator\Output\excel_output')

if __name__ == "__main__":
    create_summary()